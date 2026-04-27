from __future__ import annotations

import re
from pathlib import Path
from typing import Any

import pandas as pd
from openpyxl import load_workbook
from openpyxl.workbook.workbook import Workbook
from openpyxl.worksheet.worksheet import Worksheet


EXTERNAL_LINK_RE = re.compile(r"\[[^\]]+\]")
BROKEN_REF_TOKEN = "#REF!"


def load_workbook_safe(path: str) -> Workbook:
    """Load a workbook while preserving formulas and handling messy metadata.

    The loader attempts several mode combinations because legacy files can contain
    malformed names, stale external links, or other workbook-level artifacts.
    """

    workbook_path = Path(path)
    if not workbook_path.exists():
        raise FileNotFoundError(f"Workbook not found: {path}")

    last_error: Exception | None = None
    attempts = [
        {"data_only": False, "read_only": False, "keep_vba": True, "keep_links": True},
        {"data_only": False, "read_only": False, "keep_vba": False, "keep_links": True},
        {"data_only": False, "read_only": True, "keep_vba": False, "keep_links": True},
        {"data_only": False, "read_only": True, "keep_vba": False, "keep_links": False},
    ]

    for kwargs in attempts:
        try:
            wb = load_workbook(workbook_path, **kwargs)
            # Touch defined names defensively to flush malformed-name errors early.
            try:
                _ = list(wb.defined_names.items())
            except Exception:
                # Ignore malformed defined names; sheet-level parsing can continue.
                pass
            return wb
        except Exception as exc:  # pragma: no cover - exercised only with corrupted files
            last_error = exc

    raise RuntimeError(f"Unable to load workbook safely: {path}") from last_error


def list_sheet_names(wb: Workbook) -> list[str]:
    return list(wb.sheetnames)


def classify_sheet_name(sheet_name: str) -> str:
    name = sheet_name.strip()
    lower = name.lower()

    if name.startswith("Budget-"):
        return "budget_entity"
    if name.startswith("Actual-"):
        return "actual_entity"
    if name.startswith("CF-"):
        return "cashflow_entity"

    if any(token in lower for token in ["flash report", "汇总", "summary"]):
        return "summary"
    if any(token in lower for token in ["personnel-cn", "personnel", "people"]):
        return "driver_people"
    if any(token in lower for token in ["t&e", "travel"]):
        return "driver_travel"
    if any(token in lower for token in ["prof. fee", "professional fee", "prof fee"]):
        return "driver_prof_fee"
    if any(token in lower for token in ["other exp", "other expense"]):
        return "driver_other_exp"
    if any(token in lower for token in ["medical exp calc", "medical"]):
        return "driver_medical"
    if any(token in lower for token in ["fx lookup", "fx"]):
        return "fx"

    return "miscellaneous"


def _formula_rows(ws: Worksheet) -> list[dict[str, str]]:
    rows: list[dict[str, str]] = []
    for row in ws.iter_rows(values_only=False):
        for cell in row:
            value = cell.value
            if isinstance(value, str) and value.startswith("="):
                rows.append({"sheet_name": ws.title, "cell": cell.coordinate, "formula": value})
    return rows


def extract_formula_cells(ws: Worksheet) -> pd.DataFrame:
    return pd.DataFrame(_formula_rows(ws), columns=["sheet_name", "cell", "formula"])


def detect_external_link_formulas(ws: Worksheet) -> pd.DataFrame:
    formulas = extract_formula_cells(ws)
    if formulas.empty:
        return formulas

    mask = formulas["formula"].str.contains(EXTERNAL_LINK_RE)
    return formulas.loc[mask].reset_index(drop=True)


def detect_broken_ref_formulas(ws: Worksheet) -> pd.DataFrame:
    formulas = extract_formula_cells(ws)
    if formulas.empty:
        return formulas

    mask = formulas["formula"].str.contains(BROKEN_REF_TOKEN, regex=False)
    return formulas.loc[mask].reset_index(drop=True)


def summarize_sheet_dependencies(wb: Workbook) -> pd.DataFrame:
    rows: list[dict[str, Any]] = []
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        formula_cells = extract_formula_cells(ws)
        external = detect_external_link_formulas(ws)
        broken = detect_broken_ref_formulas(ws)

        rows.append(
            {
                "sheet_name": sheet_name,
                "sheet_type": classify_sheet_name(sheet_name),
                "formula_cell_count": int(len(formula_cells)),
                "external_link_formula_count": int(len(external)),
                "broken_ref_formula_count": int(len(broken)),
            }
        )

    return pd.DataFrame(rows)


def extract_budget_entity_sheets(wb: Workbook) -> list[str]:
    return [name for name in wb.sheetnames if name.startswith("Budget-")]


def extract_actual_entity_sheets(wb: Workbook) -> list[str]:
    return [name for name in wb.sheetnames if name.startswith("Actual-")]


def extract_cashflow_sheets(wb: Workbook) -> list[str]:
    known_tokens = ["cashflow", "cash flow", "treasury", "bank", "cf-"]
    result: list[str] = []
    for name in wb.sheetnames:
        lower = name.lower()
        if name.startswith("CF-") or any(token in lower for token in known_tokens):
            result.append(name)
    return result
