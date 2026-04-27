from __future__ import annotations

from pathlib import Path

from openpyxl import Workbook

from src.migration.legacy_parser import (
    classify_sheet_name,
    detect_broken_ref_formulas,
    detect_external_link_formulas,
    load_workbook_safe,
    summarize_sheet_dependencies,
)
from src.migration.workbook_audit import export_migration_audit_report, render_audit_markdown


def _create_test_workbook(path: Path) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "Budget-ACPL"
    ws["A1"] = "=SUM(1,2)"
    ws["A2"] = "='[90]Legacy.xlsx'!B2"
    ws["A3"] = "=#REF!+1"

    ws2 = wb.create_sheet("T&E")
    ws2["B2"] = "=A1"

    wb.create_sheet("Flash Report")
    wb.save(path)


def test_classify_sheet_name_rules() -> None:
    assert classify_sheet_name("Budget-ACPL") == "budget_entity"
    assert classify_sheet_name("Actual-ACPL") == "actual_entity"
    assert classify_sheet_name("CF-HK") == "cashflow_entity"
    assert classify_sheet_name("T&E") == "driver_travel"
    assert classify_sheet_name("Medical Exp Calc.") == "driver_medical"
    assert classify_sheet_name("FX Lookup") == "fx"
    assert classify_sheet_name("Random") == "miscellaneous"


def test_detect_external_and_broken_formulas(tmp_path: Path) -> None:
    workbook_path = tmp_path / "synthetic.xlsx"
    _create_test_workbook(workbook_path)

    wb = load_workbook_safe(str(workbook_path))
    ws = wb["Budget-ACPL"]

    external = detect_external_link_formulas(ws)
    broken = detect_broken_ref_formulas(ws)

    assert len(external) == 1
    assert external.iloc[0]["cell"] == "A2"
    assert len(broken) == 1
    assert broken.iloc[0]["cell"] == "A3"


def test_dependency_summary_output_shape(tmp_path: Path) -> None:
    workbook_path = tmp_path / "synthetic.xlsx"
    _create_test_workbook(workbook_path)

    wb = load_workbook_safe(str(workbook_path))
    summary = summarize_sheet_dependencies(wb)

    expected_cols = {
        "sheet_name",
        "sheet_type",
        "formula_cell_count",
        "external_link_formula_count",
        "broken_ref_formula_count",
    }
    assert expected_cols.issubset(set(summary.columns))
    assert len(summary) == 3


def test_markdown_report_generation(tmp_path: Path) -> None:
    workbook_path = tmp_path / "synthetic.xlsx"
    _create_test_workbook(workbook_path)

    result = export_migration_audit_report(str(workbook_path))
    markdown = render_audit_markdown(result)

    assert "# Legacy Workbook Migration Audit" in markdown
    assert "## Workbook Overview" in markdown
    assert "## Migration Recommendations" in markdown
    assert "External workbook links found." in markdown
