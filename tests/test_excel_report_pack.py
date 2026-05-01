from __future__ import annotations

import pandas as pd
from openpyxl import load_workbook

from src.exports.excel_report_pack import build_excel_report_pack, safe_sheet_name


def test_excel_file_is_created(tmp_path) -> None:
    output = tmp_path / "report_pack.xlsx"
    reports = {"Income Summary": pd.DataFrame([{"month": "2026-01", "amount": 100.0}])}

    build_excel_report_pack(output, reports)

    assert output.exists()


def test_expected_sheets_exist(tmp_path) -> None:
    output = tmp_path / "report_pack.xlsx"
    reports = {
        "Income Summary": pd.DataFrame([{"amount": 100.0}]),
        "Cashflow Summary": pd.DataFrame([{"amount": 75.0}]),
        "Driver Summary": pd.DataFrame([{"driver": "x", "amount": 20.0}]),
    }

    build_excel_report_pack(output, reports)

    wb = load_workbook(output)
    assert set(wb.sheetnames) == {"Income Summary", "Cashflow Summary", "Driver Summary"}


def test_metadata_sheet_exists_when_metadata_provided(tmp_path) -> None:
    output = tmp_path / "report_pack.xlsx"
    reports = {"Income Summary": pd.DataFrame([{"amount": 100.0}])}
    metadata = {
        "scenario": "BUD2026",
        "forecast run": "RUN-1",
        "start month": "2026-01",
        "end month": "2026-12",
        "generated timestamp": "2026-05-01T00:00:00Z",
        "notes": "test",
    }

    build_excel_report_pack(output, reports, metadata=metadata)

    wb = load_workbook(output)
    assert "Metadata" in wb.sheetnames


def test_long_sheet_names_are_safely_shortened(tmp_path) -> None:
    output = tmp_path / "report_pack.xlsx"
    long_name = "Monthly Postings Detail For Entity Group Extended"
    reports = {long_name: pd.DataFrame([{"amount": 100.0}])}

    build_excel_report_pack(output, reports)

    wb = load_workbook(output)
    expected_name = safe_sheet_name(long_name)
    assert expected_name in wb.sheetnames
    assert len(expected_name) <= 31


def test_empty_dataframes_are_handled_gracefully(tmp_path) -> None:
    output = tmp_path / "report_pack.xlsx"
    reports = {"Validation Warnings": pd.DataFrame(columns=["warning_code", "message"]) }

    build_excel_report_pack(output, reports)

    wb = load_workbook(output)
    ws = wb["Validation Warnings"]
    assert ws.max_row == 1
    assert ws.max_column == 2


def test_numeric_columns_are_preserved(tmp_path) -> None:
    output = tmp_path / "report_pack.xlsx"
    reports = {
        "Treasury Rollforward": pd.DataFrame(
            [{"month": "2026-01", "opening_cash": 1000.25, "closing_cash": 1125.75}]
        )
    }

    build_excel_report_pack(output, reports)

    wb = load_workbook(output, data_only=True)
    ws = wb["Treasury Rollforward"]
    assert isinstance(ws["B2"].value, (int, float))
    assert ws["B2"].value == 1000.25
    assert ws["C2"].value == 1125.75
